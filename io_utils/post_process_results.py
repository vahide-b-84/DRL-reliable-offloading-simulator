# post_process_results.py
# - NO average_distribution.xlsx
# - Enriches each result .xlsx by adding:
#     * "Task Distribution" sheet + chart
#     * "Recovery Strategy Distribution" sheet + chart
# - Creates ONE root-level Final_Result_All.xlsx
#     * one sheet per results folder
#     * compares Avg Reward and AVG_Failure across models (line charts)

import os
import re
import zipfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference

from config.paths import RESULTS_DIR, ensure_dirs



# ----------------------------
# Helpers
# ----------------------------

def is_valid_excel(file_path: str) -> bool:
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            return "[Content_Types].xml" in archive.namelist()
    except zipfile.BadZipFile:
        return False


def safe_sheet_delete(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]


def write_df_to_sheet(ws, df: pd.DataFrame, start_row=1, start_col=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)


def add_task_dist_chart(ws, df):
    """
    df columns:
      Server_ID, Server_Type, Primary_Tasks_Percentage, Backup_Tasks_Percentage
    """
    x_labels = [f"{sid} ({stype})" for sid, stype in zip(df["Server_ID"], df["Server_Type"])]
    ws.cell(row=1, column=5, value="Server_Label")
    for i, label in enumerate(x_labels, start=2):
        ws.cell(row=i, column=5, value=label)

    categories = Reference(ws, min_col=5, min_row=2, max_row=len(df) + 1)
    data = Reference(ws, min_col=3, min_row=1, max_col=4, max_row=len(df) + 1)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Task Distribution on Servers (Last Episode)"
    chart.y_axis.title = "Percentage of Tasks"
    chart.x_axis.title = "Server ID (Type)"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.width = 22
    chart.height = 10
    ws.add_chart(chart, "H1")


def add_recovery_strategy_chart(ws, df):
    """
    df columns: Strategy, Percentage
    """
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    data = Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1)

    chart = BarChart()
    chart.type = "col"
    chart.style = 18
    chart.title = "Recovery Strategy Distribution (Last Episode)"
    chart.y_axis.title = "Percentage of Tasks"
    chart.x_axis.title = "Recovery Strategy"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.width = 18
    chart.height = 8
    ws.add_chart(chart, "E1")


def read_required_sheets(xlsx_path: str):
    xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
    servers_df = pd.read_excel(xls, sheet_name="Servers")
    tasks_df = pd.read_excel(xls, sheet_name="TaskAssignments")
    return servers_df, tasks_df


def compute_distributions(servers_df: pd.DataFrame, tasks_df: pd.DataFrame):
    """
    Returns:
      task_distribution_df, recovery_strategy_df
    """
    server_ids = sorted(servers_df["Server_ID"].astype(int).tolist())
    server_types = servers_df.set_index("Server_ID").loc[server_ids]["Server_Type"]

    last_episode = int(tasks_df["episode"].max())
    episode_df = tasks_df[tasks_df["episode"] == last_episode].copy()

    # --- Primary count ---
    primary_count = {sid: 0 for sid in server_ids}
    for sid in episode_df["Primary"].dropna().astype(int).tolist():
        if sid in primary_count:
            primary_count[sid] += 1

    # --- Backup: only those that actually started ---
    episode_backup_df = episode_df[episode_df["Backup_Start"].notna()].copy()
    backup_count = {sid: 0 for sid in server_ids}
    for sid in episode_backup_df["Backup"].dropna().astype(int).tolist():
        if sid in backup_count:
            backup_count[sid] += 1

    total_tasks = sum(primary_count.values()) + sum(backup_count.values())
    if total_tasks <= 0:
        primary_pct = [0 for _ in server_ids]
        backup_pct = [0 for _ in server_ids]
    else:
        primary_pct = [(primary_count[sid] / total_tasks) * 100 for sid in server_ids]
        backup_pct = [(backup_count[sid] / total_tasks) * 100 for sid in server_ids]

    task_distribution_df = pd.DataFrame({
        "Server_ID": server_ids,
        "Server_Type": [server_types[sid] for sid in server_ids],
        "Primary_Tasks_Percentage": primary_pct,
        "Backup_Tasks_Percentage": backup_pct
    })

    # --- Recovery strategy distribution ---
    df2 = episode_backup_df
    retry_count = len(df2[(df2["Primary"] == df2["Backup"]) & (df2["Z"] == 0)])
    recovery_block_count = len(df2[(df2["Primary"] != df2["Backup"]) & (df2["Z"] == 0)])
    first_result_count = len(df2[df2["Z"] == 1])

    total = len(df2)
    if total <= 0:
        retry_pct = recovery_block_pct = first_result_pct = 0
    else:
        retry_pct = (retry_count / total) * 100
        recovery_block_pct = (recovery_block_count / total) * 100
        first_result_pct = (first_result_count / total) * 100

    recovery_strategy_df = pd.DataFrame({
        "Strategy": ["Retry", "Recovery Block", "First Result"],
        "Percentage": [retry_pct, recovery_block_pct, first_result_pct]
    })

    return task_distribution_df, recovery_strategy_df


def process_one_result_file(xlsx_path: str):
    if not is_valid_excel(xlsx_path):
        raise ValueError("Not a valid Excel file.")

    servers_df, tasks_df = read_required_sheets(xlsx_path)
    task_dist_df, recov_df = compute_distributions(servers_df, tasks_df)

    wb = load_workbook(xlsx_path)

    safe_sheet_delete(wb, "Task Distribution")
    ws1 = wb.create_sheet("Task Distribution")
    write_df_to_sheet(ws1, task_dist_df)
    add_task_dist_chart(ws1, task_dist_df)

    safe_sheet_delete(wb, "Recovery Strategy Distribution")
    ws2 = wb.create_sheet("Recovery Strategy Distribution")
    write_df_to_sheet(ws2, recov_df)
    add_recovery_strategy_chart(ws2, recov_df)

    wb.save(xlsx_path)
    return True


# ----------------------------
# Final_Result_All builder
# ----------------------------

def model_label_from_filename(filename: str) -> str:
    # ddpg_heterogeneous_high.xlsx -> ddpg
    base = os.path.splitext(filename)[0]
    parts = base.split("_")
    return parts[0].strip().lower() if parts else base.strip().lower()


def extract_logs_and_failure(xlsx_path: str):
    """
    Reads:
      Logs: Episode, Avg Reward
      Summary: Episode/episode, AVG_Failure
    Returns:
      logs_df: Episode + AvgReward_<model>
      fail_df: Episode + AVG_Failure_<model>
    """
    xls = pd.ExcelFile(xlsx_path, engine="openpyxl")

    logs_df = pd.read_excel(xls, sheet_name="Logs")
    logs_df.columns = [str(c).strip() for c in logs_df.columns]

    if "Episode" not in logs_df.columns and "episode" in logs_df.columns:
        logs_df = logs_df.rename(columns={"episode": "Episode"})

    if "Episode" not in logs_df.columns:
        raise ValueError("Logs missing Episode column")
    if "Avg Reward" not in logs_df.columns:
        raise ValueError("Logs missing 'Avg Reward' column")

    logs_df = logs_df[["Episode", "Avg Reward"]].copy()

    sum_df = pd.read_excel(xls, sheet_name="Summary")
    sum_df.columns = [str(c).strip() for c in sum_df.columns]

    if "Episode" not in sum_df.columns and "episode" in sum_df.columns:
        sum_df = sum_df.rename(columns={"episode": "Episode"})

    if "Episode" not in sum_df.columns:
        raise ValueError("Summary missing Episode/episode column")

    if "AVG_Failure" not in sum_df.columns:
        # If not present, keep NaN (but don't crash)
        fail_df = sum_df[["Episode"]].copy()
        fail_df["AVG_Failure"] = float("nan")
    else:
        fail_df = sum_df[["Episode", "AVG_Failure"]].copy()

    return logs_df, fail_df


def build_final_result_all(root_dir: str, folder_payloads: dict):
    """
    Creates Final_Result_All.xlsx in root_dir with one sheet per results folder.

    Each sheet contains ONE wide table:
      Episode | AvgReward_<model>... | AVG_Failure_<model>...

    Two line charts are placed without overlap:
      - Reward chart at H2
      - Failure chart at H30
    """
    if not folder_payloads:
        print("[WARN] No folders to summarize.")
        return

    out_path = os.path.join(root_dir, "Final_Result_All.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    for folder_name, payload in folder_payloads.items():
        sheet_name = folder_name[:31]
        ws = wb.create_sheet(sheet_name)

        rewards_df = payload.get("rewards")
        fail_df = payload.get("fail")

        # If nothing to write
        if (rewards_df is None or rewards_df.empty) and (fail_df is None or fail_df.empty):
            ws["A1"] = "No data"
            continue

        # ----------------------------
        # Build ONE merged wide table
        # ----------------------------
        # rewards_df: Episode + AvgReward_model cols
        # fail_df:   Episode + AVG_Failure_model cols
        merged = None

        if rewards_df is not None and not rewards_df.empty:
            merged = rewards_df.copy()

        if fail_df is not None and not fail_df.empty:
            if merged is None:
                merged = fail_df.copy()
            else:
                merged = pd.merge(merged, fail_df, on="Episode", how="outer")

        if merged is None or merged.empty:
            ws["A1"] = "No merged data"
            continue

        merged = merged.sort_values("Episode")

        # Reorder columns: Episode | AvgReward_* | AVG_Failure_*
        cols = list(merged.columns)
        ep = ["Episode"] if "Episode" in cols else []
        avg_cols = sorted([c for c in cols if c.lower().startswith("avgreward_")])
        fail_cols = sorted([c for c in cols if c.lower().startswith("avg_failure_")])
        rest = [c for c in cols if c not in ep + avg_cols + fail_cols]
        merged = merged[ep + avg_cols + fail_cols + rest]

        # Write table from A1
        write_df_to_sheet(ws, merged, start_row=1, start_col=1)

        max_row = 1 + merged.shape[0]  # header + data rows
        max_col = merged.shape[1]

        # Find column indices
        header = list(merged.columns)
        def col_index(name):
            return header.index(name) + 1  # 1-based

        c_episode = col_index("Episode") if "Episode" in header else None

        # ----------------------------
        # Chart 1: Avg Reward (Compare Models)
        # ----------------------------
        reward_cols = [c for c in header if c.lower().startswith("avgreward_")]
        if c_episode and reward_cols and max_row >= 2:
            min_c = col_index(reward_cols[0])
            max_c = col_index(reward_cols[-1])

            reward_chart = LineChart()
            reward_chart.title = "Avg Reward Over Episodes (Compare Models)"
            reward_chart.y_axis.title = "Avg Reward"
            reward_chart.x_axis.title = "Episode"

            data = Reference(ws, min_col=min_c, min_row=1, max_col=max_c, max_row=max_row)
            cats = Reference(ws, min_col=c_episode, min_row=2, max_row=max_row)

            reward_chart.add_data(data, titles_from_data=True)
            reward_chart.set_categories(cats)
            reward_chart.width = 26
            reward_chart.height = 12

            ws.add_chart(reward_chart, "H2")

        # ----------------------------
        # Chart 2: AVG_Failure (Compare Models) - place lower to avoid overlap
        # ----------------------------
        failure_cols = [c for c in header if c.lower().startswith("avg_failure_")]
        if c_episode and failure_cols and max_row >= 2:
            min_c = col_index(failure_cols[0])
            max_c = col_index(failure_cols[-1])

            fail_chart = LineChart()
            fail_chart.title = "AVG_Failure Over Episodes (Compare Models)"
            fail_chart.y_axis.title = "AVG_Failure"
            fail_chart.x_axis.title = "Episode"

            data = Reference(ws, min_col=min_c, min_row=1, max_col=max_c, max_row=max_row)
            cats = Reference(ws, min_col=c_episode, min_row=2, max_row=max_row)

            fail_chart.add_data(data, titles_from_data=True)
            fail_chart.set_categories(cats)
            fail_chart.width = 26
            fail_chart.height = 12

            # Put second chart lower (you suggested row ~30)
            ws.add_chart(fail_chart, "H30")

    wb.save(out_path)
    print(f"[OK] Created root summary: {out_path}")

# ----------------------------
# Folder scanning
# ----------------------------

def is_results_folder(name: str) -> bool:
    return bool(re.fullmatch(r"(homogeneous|heterogeneous)_(low|med|high)_results", name))


def is_result_xlsx(name: str) -> bool:
    return (
        name.lower().endswith(".xlsx")
        and not name.startswith("~$")
        and name.lower() != "final_result.xlsx"
        and name.lower() != "final_result_all.xlsx"
    )


def process_all_results(root_dir: str):
    folder_payloads = {}

    for name in os.listdir(root_dir):
        folder_path = os.path.join(root_dir, name)
        if not os.path.isdir(folder_path):
            continue
        if not is_results_folder(name):
            continue

        print(f"\n=== Processing folder: {name} ===")

        xlsx_files = [fn for fn in os.listdir(folder_path) if is_result_xlsx(fn)]
        if not xlsx_files:
            print("  [WARN] No xlsx files found.")
            continue

        # 1) Enrich each result xlsx (add dist sheets)
        for fn in xlsx_files:
            xlsx_path = os.path.join(folder_path, fn)
            try:
                process_one_result_file(xlsx_path)
                print(f"  [OK] enriched: {fn}")
            except Exception as e:
                print(f"  [FAIL] {fn} -> {e}")

        # 2) Build merged comparison frames for this folder (for root Final_Result_All)
        rewards_merged = None
        fail_merged = None

        for fn in xlsx_files:
            xlsx_path = os.path.join(folder_path, fn)
            model = model_label_from_filename(fn)

            try:
                logs_df, fail_df = extract_logs_and_failure(xlsx_path)
            except Exception as e:
                print(f"  [WARN] Skip in summary merge: {fn} -> {e}")
                continue

            logs_df = logs_df.rename(columns={"Avg Reward": f"AvgReward_{model}"})
            fail_df = fail_df.rename(columns={"AVG_Failure": f"AVG_Failure_{model}"})

            rewards_merged = logs_df if rewards_merged is None else pd.merge(rewards_merged, logs_df, on="Episode", how="outer")
            fail_merged = fail_df if fail_merged is None else pd.merge(fail_merged, fail_df, on="Episode", how="outer")

        if rewards_merged is not None:
            rewards_merged = rewards_merged.sort_values("Episode")
        if fail_merged is not None:
            fail_merged = fail_merged.sort_values("Episode")

        folder_payloads[name] = {"rewards": rewards_merged, "fail": fail_merged}

    # 3) Write ONE root summary file
    build_final_result_all(root_dir, folder_payloads)

def main():
    ensure_dirs()
    root = RESULTS_DIR
    process_all_results(root)
    print("\nDONE.")

if __name__ == "__main__":
    main()

