# save_parameters_and_logs.py
# - state-based, fixed local paths only
# - No Permutation_Number
# - Reads Excel input files ONLY from data/
# - Writes results per (scenario, state, model)
# - Creates Excel-native charts (no PNG files)
# - Adds Summary.AVG_Failure (rolling mean over last 40 episodes) + ONLY line chart in Summary

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference


from config.paths import DATA_DIR, RESULTS_DIR, ensure_dirs


def save_params_and_logs(params, log_data, task_Assignments_info):
    # Always write/read relative to project_root, not cwd, not this script's folder.
    ensure_dirs()

    scenario = getattr(params, "SCENARIO_TYPE", "heterogeneous")
    failure_state = getattr(params, "FAILURE_STATE", "high")
    model_name = str(getattr(params, "model_summary", "model")).strip().lower()

    # ---------------------------
    # Results folder + filename
    # ---------------------------
    results_dir = os.path.join(RESULTS_DIR, f"{scenario}_{failure_state}_results")
    os.makedirs(results_dir, exist_ok=True)

    filename = os.path.join(results_dir, f"{model_name}_{scenario}_{failure_state}.xlsx")

    # ---------------------------
    # Load Servers (from data/)
    # ---------------------------
    servers_xlsx_name = "homogeneous_server_info.xlsx" if scenario == "homogeneous" else "heterogeneous_server_info.xlsx"
    servers_path = os.path.join(DATA_DIR, servers_xlsx_name)
    if not os.path.exists(servers_path):
        raise FileNotFoundError(f"File not found: {servers_path}")

    server_sheet = f"{scenario.capitalize()}_state_{failure_state}"
    try:
        server_info = pd.read_excel(servers_path, sheet_name=server_sheet)
    except ValueError as e:
        raise ValueError(
            f"Sheet '{server_sheet}' not found in '{servers_path}'. "
            f"Check your generator output sheet names."
        ) from e

    # ---------------------------
    # Load Tasks (from data/)
    # ---------------------------
    task_path = os.path.join(DATA_DIR, "task_parameters.xlsx")
    if not os.path.exists(task_path):
        raise FileNotFoundError(f"File not found: {task_path}")
    task_df = pd.read_excel(task_path)

    # ---------------------------
    # Params dataframe
    # ---------------------------
    params_data = {attr: [value] for attr, value in vars(params).items()}
    df_params = pd.DataFrame(params_data).transpose().reset_index()
    df_params.columns = ["Parameter", "Value"]

    # ---------------------------
    # Logs dataframe
    # expected: (episode, avg_reward, episodic_reward, avg_delay)
    # ---------------------------
    logs_rows = []
    for log in log_data:
        logs_rows.append({
            "Episode": log[0],
            "Avg Reward": log[1] if len(log) > 1 else None,
            "Episode Reward": log[2] if len(log) > 2 else None,
            "Avg Delay": log[3] if len(log) > 3 else None,
        })
    df_logs = pd.DataFrame(logs_rows)

    # ---------------------------
    # TaskAssignments dataframe
    # ---------------------------
    df_task_Assignments = pd.DataFrame(
        task_Assignments_info,
        columns=[
            "episode",
            "task_id",
            "Primary",
            "Primary_Start",
            "Primary_End",
            "Primary_Status",
            "Backup",
            "Backup_Start",
            "Backup_End",
            "Backup_Status",
            "Z",
        ],
    )

    if not df_task_Assignments.empty:
        df_task_Assignments["Final_status"] = df_task_Assignments.apply(
            lambda row: "failure"
            if row["Primary_Status"] == "failure" and row["Backup_Status"] == "failure"
            else "success",
            axis=1,
        )
    else:
        df_task_Assignments["Final_status"] = []

    # ---------------------------
    # Summary: counts per episode + AVG_Failure (rolling mean 40)
    # ---------------------------
    if not df_task_Assignments.empty:
        summary_df = df_task_Assignments.groupby(["episode", "Final_status"]).size().unstack(fill_value=0)
    else:
        summary_df = pd.DataFrame()

    if "failure" not in summary_df.columns:
        summary_df["failure"] = 0
    if "success" not in summary_df.columns:
        summary_df["success"] = 0

    summary_df = summary_df.rename(columns={"failure": "Failure", "success": "Success"}).reset_index()

    # sort for rolling mean
    if not summary_df.empty and "episode" in summary_df.columns:
        summary_df = summary_df.sort_values("episode").reset_index(drop=True)

    if not summary_df.empty:
        summary_df["AVG_Failure"] = summary_df["Failure"].rolling(window=40, min_periods=1).mean()
    else:
        summary_df["AVG_Failure"] = []

    # ---------------------------
    # Write Excel
    # ---------------------------
    with pd.ExcelWriter(filename) as writer:
        df_params.to_excel(writer, sheet_name="Params", index=False)
        task_df.to_excel(writer, sheet_name="Tasks", index=False)
        server_info.to_excel(writer, sheet_name="Servers", index=False)
        df_logs.to_excel(writer, sheet_name="Logs", index=False)
        df_task_Assignments.to_excel(writer, sheet_name="TaskAssignments", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # ---------------------------
    # Add Excel-native charts
    # ---------------------------
    wb = load_workbook(filename)

    # Summary: ONLY line chart for AVG_Failure
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]

        max_row = summary_df.shape[0] + 1  # header included
        # columns: episode | Failure | Success | AVG_Failure
        if max_row >= 2:
            line = LineChart()
            line.title = "Average Failure Over Episodes (Rolling 40)"
            line.x_axis.title = "Episode"
            line.y_axis.title = "AVG_Failure"

            line_data = Reference(ws_sum, min_col=4, min_row=1, max_col=4, max_row=max_row)  # AVG_Failure
            line_cats = Reference(ws_sum, min_col=1, min_row=2, max_row=max_row)             # episode

            line.add_data(line_data, titles_from_data=True)
            line.set_categories(line_cats)
            line.width = 22
            line.height = 10

            # جایگذاری از بالا (چون بارچارت حذف شده، بهتره همون بالا باشه)
            ws_sum.add_chart(line, "F2")

    # Logs charts (Rewards + optional Delay) - unchanged
    if "Logs" in wb.sheetnames:
        ws_logs = wb["Logs"]
        header = [cell.value for cell in ws_logs[1]]

        def col_idx(name):
            try:
                return header.index(name) + 1
            except ValueError:
                return None

        c_episode = col_idx("Episode")
        c_avg_reward = col_idx("Avg Reward")
        c_ep_reward = col_idx("Episode Reward")
        c_avg_delay = col_idx("Avg Delay")

        max_row_logs = ws_logs.max_row

        # Rewards chart
        if c_episode and (c_avg_reward or c_ep_reward) and max_row_logs >= 2:
            rewards_chart = LineChart()
            rewards_chart.title = "Rewards per Episode"
            rewards_chart.y_axis.title = "Reward"
            rewards_chart.x_axis.title = "Episode"

            min_col = min([c for c in [c_avg_reward, c_ep_reward] if c is not None])
            max_col = max([c for c in [c_avg_reward, c_ep_reward] if c is not None])

            data = Reference(ws_logs, min_col=min_col, min_row=1, max_col=max_col, max_row=max_row_logs)
            cats = Reference(ws_logs, min_col=c_episode, min_row=2, max_row=max_row_logs)

            rewards_chart.add_data(data, titles_from_data=True)
            rewards_chart.set_categories(cats)

            ws_logs.add_chart(rewards_chart, "F2")

        # Delay chart
        if c_episode and c_avg_delay and max_row_logs >= 2:
            delay_chart = LineChart()
            delay_chart.title = "Avg Delay per Episode"
            delay_chart.y_axis.title = "Delay"
            delay_chart.x_axis.title = "Episode"

            data = Reference(ws_logs, min_col=c_avg_delay, min_row=1, max_col=c_avg_delay, max_row=max_row_logs)
            cats = Reference(ws_logs, min_col=c_episode, min_row=2, max_row=max_row_logs)

            delay_chart.add_data(data, titles_from_data=True)
            delay_chart.set_categories(cats)

            ws_logs.add_chart(delay_chart, "F20")

    wb.save(filename)
    print("successfully saved logs !")
